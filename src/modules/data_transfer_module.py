"""
Module de transfert de données entre fichiers Excel
Extraction et transfert de données avec configuration de champs
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
from typing import Dict, Any, Optional, List
import pandas as pd
from pathlib import Path
from datetime import datetime

from .base_module import BaseModule
from ..ui.components.tooltip import Tooltip
from ..ui.components.preview_table import PreviewTable
from ..ui.components.stat_card import StatCardGroup
from ..core.constants import COLORS
from ..utils.excel_utils import ExcelUtils
from ..utils.file_utils import FileUtils


class DataTransferModule(BaseModule):
    """
    Module de transfert de données entre fichiers Excel

    Fonctionnalités:
    - Définition de champs à extraire
    - Traitement par lots de fichiers
    - Création de feuilles formatées
    - Organisation automatique des fichiers
    """

    MODULE_ID = "data_transfer"
    MODULE_NAME = "Transfert de données"
    MODULE_DESCRIPTION = "Extrait et transfère des données entre fichiers Excel"
    MODULE_ICON = "📋"

    def __init__(self, *args, **kwargs):
        self.fields: List[Dict[str, str]] = []
        self.files: List[Path] = []
        super().__init__(*args, **kwargs)

    def _create_interface(self):
        """Crée l'interface du module"""
        main_scroll = ctk.CTkScrollableFrame(self.frame, fg_color="transparent")
        main_scroll.pack(fill="both", expand=True, padx=10, pady=10)

        left_panel = ctk.CTkFrame(main_scroll, fg_color="transparent", width=450)
        left_panel.pack(side="left", fill="y", padx=(0, 10))

        right_panel = ctk.CTkFrame(main_scroll, fg_color="transparent")
        right_panel.pack(side="right", fill="both", expand=True)

        self._create_files_section(left_panel)
        self._create_sheet_section(left_panel)
        self._create_fields_section(left_panel)
        self._create_action_section(left_panel)

        self._create_preview_section(right_panel)

    def _create_files_section(self, parent):
        """Section de sélection des fichiers"""
        section = ctk.CTkFrame(parent, fg_color=COLORS["bg_card"], corner_radius=10)
        section.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section,
            text="📁 FICHIERS À TRAITER",
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        btn_frame = ctk.CTkFrame(section, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=(0, 5))

        ctk.CTkButton(
            btn_frame,
            text="📁 Sélectionner dossier",
            command=self._select_folder,
            width=150
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame,
            text="📄 Sélectionner fichiers",
            command=self._select_files,
            width=150
        ).pack(side="left")

        self.files_label = ctk.CTkLabel(
            section,
            text="Aucun fichier sélectionné",
            text_color=COLORS["text_muted"]
        )
        self.files_label.pack(anchor="w", padx=15, pady=(5, 15))

    def _create_sheet_section(self, parent):
        """Section de configuration de la feuille"""
        section = ctk.CTkFrame(parent, fg_color=COLORS["bg_card"], corner_radius=10)
        section.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section,
            text="📑 FEUILLE SOURCE",
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        frame = ctk.CTkFrame(section, fg_color="transparent")
        frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkButton(
            frame,
            text="🔍 Charger exemple",
            command=self._load_example_file,
            width=150
        ).pack(side="left", padx=(0, 10))

        self.sheet_combo = ctk.CTkComboBox(
            frame,
            values=["(Charger un fichier)"],
            state="disabled",
            width=200
        )
        self.sheet_combo.pack(side="left")
        Tooltip(self.sheet_combo, "Feuille contenant les données à extraire")

    def _create_fields_section(self, parent):
        """Section de définition des champs"""
        section = ctk.CTkFrame(parent, fg_color=COLORS["bg_card"], corner_radius=10)
        section.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section,
            text="📝 CHAMPS À EXTRAIRE",
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Zone d'ajout de champ
        add_frame = ctk.CTkFrame(section, fg_color="transparent")
        add_frame.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkLabel(add_frame, text="Nom:", font=("Segoe UI", 10, "bold")).pack(side="left")
        self.field_name_entry = ctk.CTkEntry(add_frame, width=120, placeholder_text="Ex: Pilote")
        self.field_name_entry.pack(side="left", padx=(5, 10))

        ctk.CTkLabel(add_frame, text="Terme:", font=("Segoe UI", 10, "bold")).pack(side="left")
        self.field_term_entry = ctk.CTkEntry(add_frame, width=120, placeholder_text="Ex: Pilote opérationnel")
        self.field_term_entry.pack(side="left", padx=(5, 10))

        ctk.CTkButton(
            add_frame,
            text="➕",
            width=40,
            fg_color=COLORS["success"],
            command=self._add_field
        ).pack(side="left")

        # Liste des champs
        self.fields_list_frame = ctk.CTkScrollableFrame(section, fg_color="transparent", height=150)
        self.fields_list_frame.pack(fill="x", padx=15, pady=(0, 15))

        self._update_fields_list()

    def _create_action_section(self, parent):
        """Section des actions"""
        section = ctk.CTkFrame(parent, fg_color="transparent")
        section.pack(fill="x", pady=(0, 10))

        # Nom de la feuille de sortie
        output_frame = ctk.CTkFrame(section, fg_color=COLORS["bg_card"], corner_radius=10)
        output_frame.pack(fill="x", pady=(0, 10))

        frame = ctk.CTkFrame(output_frame, fg_color="transparent")
        frame.pack(fill="x", padx=15, pady=15)

        ctk.CTkLabel(frame, text="Nom feuille sortie:", width=150).pack(side="left")
        self.output_sheet_entry = ctk.CTkEntry(frame, width=200)
        self.output_sheet_entry.insert(0, "Activité")
        self.output_sheet_entry.pack(side="left", padx=(10, 0))

        # Boutons
        self.process_btn = ctk.CTkButton(
            section,
            text="🚀 TRAITER LES FICHIERS",
            font=("Segoe UI", 14, "bold"),
            height=50,
            fg_color=COLORS["success"],
            command=self.start_execution
        )
        self.process_btn.pack(fill="x", pady=(0, 10))

        # Progression
        self.progress_bar = ctk.CTkProgressBar(section, height=8)
        self.progress_bar.pack(fill="x")
        self.progress_bar.set(0)

        self.status_label = ctk.CTkLabel(
            section,
            text="Prêt",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"]
        )
        self.status_label.pack(anchor="w", pady=(5, 0))

    def _create_preview_section(self, parent):
        """Section de prévisualisation"""
        # Statistiques
        self.stats_cards = StatCardGroup(parent)
        self.stats_cards.pack(fill="x", pady=(0, 10))

        self.stats_cards.add_card("total", "Fichiers", "0", "📁")
        self.stats_cards.add_card("success", "Succès", "0", "✓", color=COLORS["success"])
        self.stats_cards.add_card("errors", "Erreurs", "0", "✗", color=COLORS["error"])

        # Aperçu des données extraites
        self.preview_table = PreviewTable(
            parent,
            title="Aperçu des données extraites"
        )
        self.preview_table.pack(fill="both", expand=True)

    def _select_folder(self):
        """Sélectionne un dossier de fichiers"""
        folder = filedialog.askdirectory(title="Sélectionner le dossier")
        if folder:
            self.files = FileUtils.list_excel_files(folder)
            self._update_files_label()

    def _select_files(self):
        """Sélectionne des fichiers individuels"""
        files = filedialog.askopenfilenames(
            title="Sélectionner les fichiers",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls *.xlsm")]
        )
        if files:
            self.files = [Path(f) for f in files]
            self._update_files_label()

    def _update_files_label(self):
        """Met à jour le label du nombre de fichiers"""
        n = len(self.files)
        if n > 0:
            self.files_label.configure(
                text=f"✓ {n} fichier(s) sélectionné(s)",
                text_color=COLORS["success"]
            )
        else:
            self.files_label.configure(
                text="Aucun fichier sélectionné",
                text_color=COLORS["text_muted"]
            )

    def _load_example_file(self):
        """Charge un fichier exemple pour la configuration"""
        filepath = filedialog.askopenfilename(
            title="Sélectionner un fichier exemple",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls *.xlsm")]
        )

        if filepath:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, read_only=True)
                sheets = wb.sheetnames
                wb.close()

                self.sheet_combo.configure(state="normal", values=sheets)
                if sheets:
                    self.sheet_combo.set(sheets[0])

                self._preview_extraction(filepath)
                self.log_info(f"Fichier exemple chargé: {Path(filepath).name}")

            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de lire le fichier:\n{str(e)}")

    def _add_field(self):
        """Ajoute un champ à extraire"""
        name = self.field_name_entry.get().strip()
        term = self.field_term_entry.get().strip()

        if not name or not term:
            messagebox.showwarning("Attention", "Veuillez remplir le nom et le terme")
            return

        self.fields.append({"name": name, "term": term})
        self.field_name_entry.delete(0, "end")
        self.field_term_entry.delete(0, "end")

        self._update_fields_list()
        self.log_info(f"Champ ajouté: {name}")

    def _remove_field(self, index: int):
        """Supprime un champ"""
        if 0 <= index < len(self.fields):
            del self.fields[index]
            self._update_fields_list()

    def _update_fields_list(self):
        """Met à jour l'affichage de la liste des champs"""
        for widget in self.fields_list_frame.winfo_children():
            widget.destroy()

        if not self.fields:
            ctk.CTkLabel(
                self.fields_list_frame,
                text="Aucun champ défini. Ajoutez des champs ci-dessus.",
                text_color=COLORS["text_muted"]
            ).pack(pady=10)
            return

        for idx, field in enumerate(self.fields):
            frame = ctk.CTkFrame(self.fields_list_frame, fg_color=("gray85", "gray25"))
            frame.pack(fill="x", pady=2)

            ctk.CTkLabel(
                frame,
                text=f"{idx + 1}. {field['name']}",
                font=("Segoe UI", 11, "bold")
            ).pack(side="left", padx=10, pady=5)

            ctk.CTkLabel(
                frame,
                text=f"→ {field['term']}",
                text_color=COLORS["text_muted"]
            ).pack(side="left", padx=(0, 10))

            ctk.CTkButton(
                frame,
                text="❌",
                width=30,
                height=24,
                fg_color=COLORS["error"],
                command=lambda i=idx: self._remove_field(i)
            ).pack(side="right", padx=5, pady=3)

    def _preview_extraction(self, filepath: str):
        """Prévisualise l'extraction sur un fichier"""
        if not self.fields:
            return

        try:
            sheet_name = self.sheet_combo.get()
            data = self._extract_data_from_file(filepath, sheet_name)

            if data:
                df = pd.DataFrame([data])
                self.preview_table.load_data(df)

        except Exception as e:
            self.log_error(f"Erreur de prévisualisation: {e}")

    def _extract_data_from_file(self, filepath: str, sheet_name: str) -> Optional[Dict]:
        """Extrait les données d'un fichier avec détection intelligente"""
        from openpyxl import load_workbook

        wb = load_workbook(filepath, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None

        sheet = wb[sheet_name]
        data = {}

        for field in self.fields:
            term = field['term'].lower()
            value = None

            for row in sheet.iter_rows(min_row=1, max_row=200, values_only=False):
                for cell in row:
                    if cell.value is None:
                        continue

                    cell_text = str(cell.value).strip().lower()
                    if term in cell_text:
                        col_idx = cell.column
                        row_idx = cell.row

                        # Stratégie 1: Chercher à droite sur la même ligne
                        for next_col in range(col_idx + 1, min(col_idx + 5, sheet.max_column + 1)):
                            next_cell = sheet.cell(row=row_idx, column=next_col)
                            if next_cell.value is not None:
                                val = next_cell.value
                                if isinstance(val, datetime):
                                    val = val.strftime("%d/%m/%Y")
                                value = str(val).strip()
                                if value:
                                    break

                        # Stratégie 2: Si rien trouvé, chercher en dessous (format vertical)
                        if not value:
                            for next_row in range(row_idx + 1, min(row_idx + 3, sheet.max_row + 1)):
                                below_cell = sheet.cell(row=next_row, column=col_idx)
                                if below_cell.value is not None:
                                    val = below_cell.value
                                    if isinstance(val, datetime):
                                        val = val.strftime("%d/%m/%Y")
                                    value = str(val).strip()
                                    if value:
                                        break

                        # Stratégie 3: Chercher cellule fusionnée adjacente
                        if not value and hasattr(sheet, 'merged_cells'):
                            for merged_range in sheet.merged_cells.ranges:
                                if (merged_range.min_row == row_idx and
                                    merged_range.min_col > col_idx and
                                    merged_range.min_col <= col_idx + 3):
                                    merged_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                                    if merged_cell.value is not None:
                                        val = merged_cell.value
                                        if isinstance(val, datetime):
                                            val = val.strftime("%d/%m/%Y")
                                        value = str(val).strip()
                                        break

                        if value:
                            break
                if value:
                    break

            data[field['name']] = value or ""

        wb.close()
        return data

    def validate_inputs(self) -> tuple[bool, str]:
        """Valide les entrées"""
        if not self.files:
            return False, "Aucun fichier sélectionné"

        if not self.fields:
            return False, "Aucun champ défini"

        sheet = self.sheet_combo.get()
        if not sheet or sheet == "(Charger un fichier)":
            return False, "Veuillez sélectionner une feuille source"

        return True, ""

    def _execute_task(self) -> Dict[str, Any]:
        """Exécute le traitement des fichiers"""
        sheet_name = self.sheet_combo.get()
        output_sheet = self.output_sheet_entry.get().strip() or "Activité"

        total = len(self.files)
        success = 0
        errors = 0

        for idx, filepath in enumerate(self.files):
            if self.is_cancelled():
                break

            self.update_progress((idx + 1) / total)
            self.update_status(f"Traitement: {filepath.name}")

            try:
                data = self._extract_data_from_file(str(filepath), sheet_name)

                if data and any(data.values()):
                    self._create_activity_sheet(filepath, data, output_sheet)
                    success += 1
                    self.log_success(f"Traité: {filepath.name}")
                else:
                    errors += 1
                    self.log_warning(f"Aucune donnée trouvée: {filepath.name}")

            except Exception as e:
                errors += 1
                self.log_error(f"Erreur {filepath.name}: {e}")

        # Mise à jour de l'interface
        self.frame.after(0, lambda: self._update_stats(total, success, errors))

        return {"total": total, "success": success, "errors": errors}

    def _create_activity_sheet(self, filepath: Path, data: Dict, sheet_name: str):
        """Crée la feuille d'activité dans le fichier"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = load_workbook(filepath)

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        # Styles
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        label_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-tête
        ws.merge_cells('A1:B1')
        ws['A1'] = "DONNÉES EXTRAITES"
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Données
        row = 2
        for field in self.fields:
            ws.cell(row=row, column=1, value=field['name']).fill = label_fill
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2, value=data.get(field['name'], "")).border = border
            row += 1

        # Ajuster les colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

        wb.save(filepath)
        wb.close()

    def _update_stats(self, total: int, success: int, errors: int):
        """Met à jour les statistiques affichées"""
        self.stats_cards.update_card("total", str(total))
        self.stats_cards.update_card("success", str(success))
        self.stats_cards.update_card("errors", str(errors))

        if errors > 0:
            self.update_status(f"Terminé avec {errors} erreur(s)", "warning")
        else:
            self.update_status(f"Terminé: {success}/{total} fichiers traités", "success")

    def update_status(self, message: str, level: str = "info"):
        color_map = {
            "info": COLORS["text_muted"],
            "success": COLORS["success"],
            "warning": COLORS["warning"],
            "error": COLORS["error"]
        }
        self.status_label.configure(
            text=message,
            text_color=color_map.get(level, COLORS["text_muted"])
        )

    def update_progress(self, progress: float):
        self.progress_bar.set(progress)

    def reset(self):
        super().reset()
        self.files = []
        self._update_files_label()
        self.preview_table.clear()
        self.stats_cards.reset_all()
