"""
Module de conversion CSV et gestion de fichiers
Conversion CSV <-> Excel et fusion de fichiers
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
from typing import Dict, Any, Optional, List
import pandas as pd
from pathlib import Path

from .base_module import BaseModule
from ..ui.components.tooltip import Tooltip
from ..ui.components.preview_table import PreviewTable
from ..ui.components.stat_card import StatCardGroup
from ..core.constants import COLORS
from ..utils.excel_utils import ExcelUtils


class CSVConverterModule(BaseModule):
    """
    Module de conversion et gestion de fichiers

    Fonctionnalités:
    - Conversion CSV vers Excel
    - Conversion Excel vers CSV
    - Fusion de fichiers Excel
    - Exploration de fichiers Excel
    """

    MODULE_ID = "csv_converter"
    MODULE_NAME = "Conversion & Fusion"
    MODULE_DESCRIPTION = "Convertit et fusionne des fichiers CSV/Excel"
    MODULE_ICON = "🔄"

    def _create_interface(self):
        """Crée l'interface du module"""
        # Onglets pour les différentes fonctionnalités
        self.tabview = ctk.CTkTabview(self.frame, fg_color=COLORS["bg_card"])
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)

        self.tabview.add("CSV → Excel")
        self.tabview.add("Excel → CSV")
        self.tabview.add("Fusion Excel")
        self.tabview.add("Exploration")

        self._create_csv_to_excel_tab()
        self._create_excel_to_csv_tab()
        self._create_merge_tab()
        self._create_exploration_tab()

    def _create_csv_to_excel_tab(self):
        """Onglet conversion CSV vers Excel"""
        tab = self.tabview.tab("CSV → Excel")

        scroll = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # Section fichier CSV
        section1 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section1.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(section1, text="📄 FICHIER CSV", font=("Segoe UI", 12, "bold")).pack(
            anchor="w", padx=15, pady=(15, 10))

        frame1 = ctk.CTkFrame(section1, fg_color="transparent")
        frame1.pack(fill="x", padx=15, pady=(0, 15))

        self.csv_path_var = ctk.StringVar()
        ctk.CTkEntry(frame1, textvariable=self.csv_path_var, width=300, state="disabled").pack(
            side="left", padx=(0, 10))
        ctk.CTkButton(frame1, text="📁 Parcourir", width=100,
                      command=self._browse_csv).pack(side="left")

        # Options
        options_frame = ctk.CTkFrame(section1, fg_color="transparent")
        options_frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(options_frame, text="Séparateur:").pack(side="left")
        self.separator_entry = ctk.CTkEntry(options_frame, width=50)
        self.separator_entry.insert(0, ",")
        self.separator_entry.pack(side="left", padx=10)
        Tooltip(self.separator_entry, "Caractère de séparation (virgule, point-virgule, etc.)")

        ctk.CTkLabel(options_frame, text="Encodage:").pack(side="left", padx=(20, 0))
        self.encoding_combo = ctk.CTkComboBox(
            options_frame, values=["utf-8", "latin-1", "cp1252", "iso-8859-1"], width=100)
        self.encoding_combo.set("utf-8")
        self.encoding_combo.pack(side="left", padx=10)

        # Section fichier Excel de sortie
        section2 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section2.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(section2, text="📊 FICHIER EXCEL DE SORTIE", font=("Segoe UI", 12, "bold")).pack(
            anchor="w", padx=15, pady=(15, 10))

        frame2 = ctk.CTkFrame(section2, fg_color="transparent")
        frame2.pack(fill="x", padx=15, pady=(0, 10))

        self.excel_output_var = ctk.StringVar()
        ctk.CTkEntry(frame2, textvariable=self.excel_output_var, width=300, state="disabled").pack(
            side="left", padx=(0, 10))
        ctk.CTkButton(frame2, text="📁 Parcourir", width=100,
                      command=self._browse_excel_output).pack(side="left")

        frame3 = ctk.CTkFrame(section2, fg_color="transparent")
        frame3.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(frame3, text="Nom de la feuille:").pack(side="left")
        self.sheet_name_entry = ctk.CTkEntry(frame3, width=200)
        self.sheet_name_entry.insert(0, "Données")
        self.sheet_name_entry.pack(side="left", padx=10)

        # Bouton de conversion
        ctk.CTkButton(
            scroll,
            text="🔄 CONVERTIR CSV → EXCEL",
            font=("Segoe UI", 14, "bold"),
            height=50,
            fg_color=COLORS["success"],
            command=self._convert_csv_to_excel
        ).pack(fill="x", pady=10)

    def _create_excel_to_csv_tab(self):
        """Onglet conversion Excel vers CSV"""
        tab = self.tabview.tab("Excel → CSV")

        scroll = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # Section fichier Excel
        section1 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section1.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(section1, text="📊 FICHIER EXCEL", font=("Segoe UI", 12, "bold")).pack(
            anchor="w", padx=15, pady=(15, 10))

        frame1 = ctk.CTkFrame(section1, fg_color="transparent")
        frame1.pack(fill="x", padx=15, pady=(0, 10))

        self.excel_input_var = ctk.StringVar()
        ctk.CTkEntry(frame1, textvariable=self.excel_input_var, width=300, state="disabled").pack(
            side="left", padx=(0, 10))
        ctk.CTkButton(frame1, text="📁 Parcourir", width=100,
                      command=self._browse_excel_input).pack(side="left")

        frame2 = ctk.CTkFrame(section1, fg_color="transparent")
        frame2.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(frame2, text="Feuille:").pack(side="left")
        self.sheet_combo = ctk.CTkComboBox(frame2, values=["(Charger un fichier)"], width=200,
                                           state="disabled")
        self.sheet_combo.pack(side="left", padx=10)

        # Options
        options_frame = ctk.CTkFrame(section1, fg_color="transparent")
        options_frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(options_frame, text="Séparateur:").pack(side="left")
        self.csv_separator_entry = ctk.CTkEntry(options_frame, width=50)
        self.csv_separator_entry.insert(0, ",")
        self.csv_separator_entry.pack(side="left", padx=10)

        # Bouton de conversion
        ctk.CTkButton(
            scroll,
            text="🔄 CONVERTIR EXCEL → CSV",
            font=("Segoe UI", 14, "bold"),
            height=50,
            fg_color=COLORS["success"],
            command=self._convert_excel_to_csv
        ).pack(fill="x", pady=10)

    def _create_merge_tab(self):
        """Onglet fusion de fichiers"""
        tab = self.tabview.tab("Fusion Excel")

        scroll = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # Section fichiers à fusionner
        section1 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section1.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(section1, text="📁 FICHIERS À FUSIONNER", font=("Segoe UI", 12, "bold")).pack(
            anchor="w", padx=15, pady=(15, 10))

        frame1 = ctk.CTkFrame(section1, fg_color="transparent")
        frame1.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkButton(frame1, text="📁 Sélectionner fichiers", width=150,
                      command=self._select_merge_files).pack(side="left")

        self.merge_files_label = ctk.CTkLabel(frame1, text="Aucun fichier", text_color=COLORS["text_muted"])
        self.merge_files_label.pack(side="left", padx=15)

        self.merge_files_list: List[str] = []

        # Options
        options_frame = ctk.CTkFrame(section1, fg_color="transparent")
        options_frame.pack(fill="x", padx=15, pady=(0, 15))

        self.skip_headers_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(options_frame, text="Ignorer les en-têtes (sauf premier fichier)",
                        variable=self.skip_headers_var).pack(anchor="w")

        # Bouton de fusion
        ctk.CTkButton(
            scroll,
            text="🔗 FUSIONNER LES FICHIERS",
            font=("Segoe UI", 14, "bold"),
            height=50,
            fg_color=COLORS["success"],
            command=self._merge_files
        ).pack(fill="x", pady=10)

    def _create_exploration_tab(self):
        """Onglet exploration de fichiers"""
        tab = self.tabview.tab("Exploration")

        scroll = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # Section fichier
        section1 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section1.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(section1, text="📊 FICHIER À EXPLORER", font=("Segoe UI", 12, "bold")).pack(
            anchor="w", padx=15, pady=(15, 10))

        frame1 = ctk.CTkFrame(section1, fg_color="transparent")
        frame1.pack(fill="x", padx=15, pady=(0, 15))

        self.explore_path_var = ctk.StringVar()
        ctk.CTkEntry(frame1, textvariable=self.explore_path_var, width=300, state="disabled").pack(
            side="left", padx=(0, 10))
        ctk.CTkButton(frame1, text="📁 Parcourir", width=100,
                      command=self._browse_explore).pack(side="left")

        # Actions
        actions_frame = ctk.CTkFrame(section1, fg_color="transparent")
        actions_frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkButton(actions_frame, text="📋 Lister feuilles", width=150,
                      command=self._list_sheets).pack(side="left", padx=(0, 10))
        ctk.CTkButton(actions_frame, text="📊 Lister colonnes", width=150,
                      command=self._list_columns).pack(side="left")

        # Zone de résultats
        self.explore_text = ctk.CTkTextbox(scroll, height=300, font=("Consolas", 11))
        self.explore_text.pack(fill="both", expand=True, pady=10)

    # === Méthodes CSV -> Excel ===

    def _browse_csv(self):
        filepath = filedialog.askopenfilename(
            title="Sélectionner un fichier CSV",
            filetypes=[("Fichiers CSV", "*.csv *.txt")]
        )
        if filepath:
            self.csv_path_var.set(filepath)

    def _browse_excel_output(self):
        filepath = filedialog.asksaveasfilename(
            title="Enregistrer sous",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")]
        )
        if filepath:
            self.excel_output_var.set(filepath)

    def _convert_csv_to_excel(self):
        csv_path = self.csv_path_var.get()
        excel_path = self.excel_output_var.get()
        separator = self.separator_entry.get() or ","
        encoding = self.encoding_combo.get()
        sheet_name = self.sheet_name_entry.get().strip() or "Données"

        if not csv_path or not excel_path:
            messagebox.showwarning("Attention", "Veuillez sélectionner les fichiers")
            return

        try:
            df = pd.read_csv(csv_path, sep=separator, encoding=encoding)

            success, error = ExcelUtils.write_dataframe_to_excel(df, excel_path, sheet_name)

            if success:
                messagebox.showinfo("Succès", f"Conversion terminée!\n{len(df)} lignes exportées")
                self.log_success(f"CSV converti: {csv_path}")
            else:
                messagebox.showerror("Erreur", error)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.log_error(str(e))

    # === Méthodes Excel -> CSV ===

    def _browse_excel_input(self):
        filepath = filedialog.askopenfilename(
            title="Sélectionner un fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls *.xlsm")]
        )
        if filepath:
            self.excel_input_var.set(filepath)
            self._load_excel_sheets(filepath)

    def _load_excel_sheets(self, filepath: str):
        try:
            df_dict = pd.read_excel(filepath, sheet_name=None, nrows=0)
            sheets = list(df_dict.keys())
            self.sheet_combo.configure(state="normal", values=sheets)
            if sheets:
                self.sheet_combo.set(sheets[0])
        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def _convert_excel_to_csv(self):
        excel_path = self.excel_input_var.get()
        sheet = self.sheet_combo.get()
        separator = self.csv_separator_entry.get() or ","

        if not excel_path:
            messagebox.showwarning("Attention", "Veuillez sélectionner un fichier")
            return

        csv_path = filedialog.asksaveasfilename(
            title="Enregistrer le CSV",
            defaultextension=".csv",
            filetypes=[("Fichiers CSV", "*.csv")]
        )

        if csv_path:
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet)
                df.to_csv(csv_path, sep=separator, index=False, encoding='utf-8')

                messagebox.showinfo("Succès", f"Conversion terminée!\n{len(df)} lignes exportées")
                self.log_success(f"Excel converti en CSV: {csv_path}")

            except Exception as e:
                messagebox.showerror("Erreur", str(e))
                self.log_error(str(e))

    # === Méthodes Fusion ===

    def _select_merge_files(self):
        files = filedialog.askopenfilenames(
            title="Sélectionner les fichiers à fusionner",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls *.xlsm")]
        )
        if files:
            self.merge_files_list = list(files)
            self.merge_files_label.configure(
                text=f"{len(files)} fichier(s) sélectionné(s)",
                text_color=COLORS["success"]
            )

    def _merge_files(self):
        if not self.merge_files_list:
            messagebox.showwarning("Attention", "Veuillez sélectionner des fichiers")
            return

        output_path = filedialog.asksaveasfilename(
            title="Enregistrer le fichier fusionné",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")]
        )

        if output_path:
            try:
                success, rows, error = ExcelUtils.merge_excel_files(
                    self.merge_files_list,
                    output_path,
                    skip_headers=self.skip_headers_var.get()
                )

                if success:
                    messagebox.showinfo("Succès", f"Fusion terminée!\n{rows} lignes dans le fichier fusionné")
                    self.log_success(f"Fichiers fusionnés: {output_path}")
                else:
                    messagebox.showerror("Erreur", error)

            except Exception as e:
                messagebox.showerror("Erreur", str(e))
                self.log_error(str(e))

    # === Méthodes Exploration ===

    def _browse_explore(self):
        filepath = filedialog.askopenfilename(
            title="Sélectionner un fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls *.xlsm")]
        )
        if filepath:
            self.explore_path_var.set(filepath)

    def _list_sheets(self):
        filepath = self.explore_path_var.get()
        if not filepath:
            messagebox.showwarning("Attention", "Veuillez sélectionner un fichier")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            sheets = wb.sheetnames
            wb.close()

            result = f"FEUILLES DU FICHIER\n{'=' * 40}\n\n"
            result += f"Fichier: {Path(filepath).name}\n"
            result += f"Nombre de feuilles: {len(sheets)}\n\n"

            for i, sheet in enumerate(sheets, 1):
                result += f"  {i}. {sheet}\n"

            self.explore_text.delete("1.0", "end")
            self.explore_text.insert("1.0", result)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def _list_columns(self):
        filepath = self.explore_path_var.get()
        if not filepath:
            messagebox.showwarning("Attention", "Veuillez sélectionner un fichier")
            return

        try:
            df = pd.read_excel(filepath, nrows=0)
            columns = list(df.columns)

            result = f"COLONNES DU FICHIER\n{'=' * 40}\n\n"
            result += f"Fichier: {Path(filepath).name}\n"
            result += f"Nombre de colonnes: {len(columns)}\n\n"

            for i, col in enumerate(columns, 1):
                result += f"  {i}. {col}\n"

            self.explore_text.delete("1.0", "end")
            self.explore_text.insert("1.0", result)

        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    # === Méthodes obligatoires de BaseModule ===

    def validate_inputs(self) -> tuple[bool, str]:
        return True, ""

    def _execute_task(self) -> Dict[str, Any]:
        return {}
