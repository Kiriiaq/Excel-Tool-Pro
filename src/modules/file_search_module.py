"""
Module de recherche et filtrage de fichiers Excel
Recherche avancée dans les données avec multiple critères
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
from typing import Dict, Any, Optional, List
import pandas as pd
import re
import threading

from .base_module import BaseModule
from ..ui.components.tooltip import Tooltip
from ..ui.components.file_selector import FileSelector
from ..ui.components.preview_table import PreviewTable
from ..ui.components.stat_card import StatCardGroup
from ..core.constants import COLORS
from ..utils.excel_utils import ExcelUtils


class FileSearchModule(BaseModule):
    """
    Module de recherche avancée dans les fichiers Excel

    Fonctionnalités:
    - Recherche par mots-clés avec plusieurs modes
    - Recherche depuis une liste de mots (fichier ou texte)
    - Opérateurs logiques (AND, OR)
    - Correspondance approximative (fuzzy matching)
    - Surbrillance des correspondances
    - Filtres par colonnes
    - Export des résultats avec statistiques
    """

    MODULE_ID = "file_search"
    MODULE_NAME = "Recherche de données"
    MODULE_DESCRIPTION = "Recherche avancée dans les fichiers Excel"
    MODULE_ICON = "🔍"

    def __init__(self, *args, **kwargs):
        self._stop_event = threading.Event()
        super().__init__(*args, **kwargs)

    def _create_interface(self):
        """Crée l'interface du module"""
        # Onglets pour différents modes de recherche
        self.tabview = ctk.CTkTabview(self.frame, fg_color=COLORS["bg_card"])
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)

        self.tabview.add("Recherche simple")
        self.tabview.add("Recherche par liste")
        self.tabview.add("Résultats")

        self._create_simple_search_tab()
        self._create_list_search_tab()
        self._create_results_tab()

        self.df_results: Optional[pd.DataFrame] = None
        self.match_details: List[Dict] = []

    def _create_simple_search_tab(self):
        """Onglet de recherche simple"""
        tab = self.tabview.tab("Recherche simple")

        scroll = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # === Section Fichier ===
        section1 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section1.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section1, text="📁 FICHIER SOURCE", font=("Segoe UI", 12, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        self.file_selector = FileSelector(
            section1,
            label="Fichier Excel",
            tooltip="Sélectionnez le fichier à analyser",
            on_file_loaded=self._on_file_loaded
        )
        self.file_selector.pack(fill="x", padx=15, pady=(0, 15))

        # === Section Recherche ===
        section2 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section2.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section2, text="🔍 RECHERCHE", font=("Segoe UI", 12, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Champ de recherche
        search_frame = ctk.CTkFrame(section2, fg_color="transparent")
        search_frame.pack(fill="x", padx=15, pady=(0, 10))

        lbl = ctk.CTkLabel(search_frame, text="Mots-clés:", width=100)
        lbl.pack(side="left")
        Tooltip(lbl, "Entrez les termes à rechercher, séparés par des virgules")

        self.search_entry = ctk.CTkEntry(
            search_frame, placeholder_text="terme1, terme2...", width=300
        )
        self.search_entry.pack(side="left", padx=(10, 0))

        # Mode de recherche
        mode_frame = ctk.CTkFrame(section2, fg_color="transparent")
        mode_frame.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkLabel(mode_frame, text="Mode:", width=100).pack(side="left")

        self.search_mode_var = ctk.StringVar(value="contains")
        modes = [
            ("Contient", "contains"),
            ("Mot exact", "exact"),
            ("Commence par", "starts"),
            ("Finit par", "ends"),
            ("Regex", "regex"),
            ("Approx.", "fuzzy")
        ]

        for text, value in modes:
            rb = ctk.CTkRadioButton(
                mode_frame, text=text, variable=self.search_mode_var,
                value=value, font=("Segoe UI", 10)
            )
            rb.pack(side="left", padx=5)

        # Options
        options_frame = ctk.CTkFrame(section2, fg_color="transparent")
        options_frame.pack(fill="x", padx=15, pady=(0, 10))

        self.case_sensitive_var = ctk.BooleanVar(value=False)
        cb1 = ctk.CTkCheckBox(
            options_frame, text="Sensible à la casse", variable=self.case_sensitive_var
        )
        cb1.pack(side="left", padx=(0, 15))

        self.and_mode_var = ctk.BooleanVar(value=False)
        cb2 = ctk.CTkCheckBox(
            options_frame, text="Mode AND", variable=self.and_mode_var
        )
        cb2.pack(side="left", padx=(0, 15))
        Tooltip(cb2, "Tous les termes doivent être présents")

        self.highlight_var = ctk.BooleanVar(value=True)
        cb3 = ctk.CTkCheckBox(
            options_frame, text="Surbrillance", variable=self.highlight_var
        )
        cb3.pack(side="left")
        Tooltip(cb3, "Surligner les correspondances dans l'export")

        # Seuil fuzzy
        fuzzy_frame = ctk.CTkFrame(section2, fg_color="transparent")
        fuzzy_frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(fuzzy_frame, text="Seuil similarité:", width=100).pack(side="left")
        self.fuzzy_slider = ctk.CTkSlider(
            fuzzy_frame, from_=50, to=100, number_of_steps=50, width=150
        )
        self.fuzzy_slider.set(80)
        self.fuzzy_slider.pack(side="left", padx=(10, 5))

        self.fuzzy_label = ctk.CTkLabel(fuzzy_frame, text="80%", width=40)
        self.fuzzy_label.pack(side="left")

        self.fuzzy_slider.configure(command=self._update_fuzzy_label)

        # === Section Filtres ===
        section3 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section3.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section3, text="🎯 COLONNES", font=("Segoe UI", 12, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Boutons de sélection
        btn_frame = ctk.CTkFrame(section3, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=(0, 5))

        ctk.CTkButton(
            btn_frame, text="Tout sélectionner", width=120,
            command=lambda: self._select_all_columns(True)
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="Tout désélectionner", width=120,
            command=lambda: self._select_all_columns(False)
        ).pack(side="left")

        # Liste des colonnes
        self.columns_frame = ctk.CTkScrollableFrame(
            section3, fg_color="transparent", height=100
        )
        self.columns_frame.pack(fill="x", padx=15, pady=(0, 15))

        self.column_checkboxes = {}

        # === Section Actions ===
        action_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        action_frame.pack(fill="x", pady=(0, 10))

        self.search_btn = ctk.CTkButton(
            action_frame,
            text="🔍 RECHERCHER",
            font=("Segoe UI", 14, "bold"),
            height=50,
            fg_color=COLORS["accent_primary"],
            command=self._start_simple_search
        )
        self.search_btn.pack(fill="x", pady=(0, 10))

        # Progression
        self.progress_bar = ctk.CTkProgressBar(action_frame, height=8)
        self.progress_bar.pack(fill="x")
        self.progress_bar.set(0)

        self.status_label = ctk.CTkLabel(
            action_frame, text="Prêt", font=("Segoe UI", 10), text_color=COLORS["text_muted"]
        )
        self.status_label.pack(anchor="w", pady=(5, 0))

    def _create_list_search_tab(self):
        """Onglet de recherche par liste de mots"""
        tab = self.tabview.tab("Recherche par liste")

        scroll = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # === Section Fichier Excel ===
        section1 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section1.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section1, text="📊 FICHIER EXCEL À ANALYSER", font=("Segoe UI", 12, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        self.list_file_selector = FileSelector(
            section1,
            label="Fichier Excel",
            tooltip="Fichier dans lequel effectuer la recherche",
            on_file_loaded=self._on_list_file_loaded
        )
        self.list_file_selector.pack(fill="x", padx=15, pady=(0, 10))

        col_frame = ctk.CTkFrame(section1, fg_color="transparent")
        col_frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(col_frame, text="Colonne de recherche:", width=140).pack(side="left")
        self.search_col_combo = ctk.CTkComboBox(
            col_frame, values=["(Charger fichier)"], state="disabled", width=200
        )
        self.search_col_combo.pack(side="left", padx=(10, 0))

        # === Section Liste de mots ===
        section2 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section2.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section2, text="📝 LISTE DE MOTS À RECHERCHER", font=("Segoe UI", 12, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Mode d'entrée
        mode_frame = ctk.CTkFrame(section2, fg_color="transparent")
        mode_frame.pack(fill="x", padx=15, pady=(0, 10))

        self.list_mode_var = ctk.StringVar(value="text")
        ctk.CTkRadioButton(
            mode_frame, text="Saisie manuelle", variable=self.list_mode_var,
            value="text", command=self._toggle_list_mode
        ).pack(side="left", padx=(0, 20))

        ctk.CTkRadioButton(
            mode_frame, text="Depuis fichier", variable=self.list_mode_var,
            value="file", command=self._toggle_list_mode
        ).pack(side="left")

        # Zone de texte pour saisie manuelle
        self.words_text = ctk.CTkTextbox(
            section2, height=100, font=("Consolas", 10)
        )
        self.words_text.pack(fill="x", padx=15, pady=(0, 10))
        self.words_text.insert("1.0", "# Un mot par ligne\n# Lignes commençant par # ignorées")

        # Sélection fichier
        self.words_file_frame = ctk.CTkFrame(section2, fg_color="transparent")
        self.words_file_frame.pack(fill="x", padx=15, pady=(0, 15))
        self.words_file_frame.pack_forget()  # Caché par défaut

        self.words_file_var = ctk.StringVar()
        ctk.CTkEntry(
            self.words_file_frame, textvariable=self.words_file_var, width=300, state="disabled"
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            self.words_file_frame, text="📁 Parcourir", width=100,
            command=self._browse_words_file
        ).pack(side="left")

        # === Section Options ===
        section3 = ctk.CTkFrame(scroll, fg_color=COLORS["bg_card"], corner_radius=10)
        section3.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section3, text="⚙️ OPTIONS", font=("Segoe UI", 12, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        options_frame = ctk.CTkFrame(section3, fg_color="transparent")
        options_frame.pack(fill="x", padx=15, pady=(0, 10))

        self.list_case_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            options_frame, text="Sensible à la casse", variable=self.list_case_var
        ).pack(side="left", padx=(0, 15))

        self.list_exact_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            options_frame, text="Mot exact", variable=self.list_exact_var
        ).pack(side="left", padx=(0, 15))
        Tooltip(options_frame.winfo_children()[-1], "Correspond au mot entier (avec \\b)")

        self.list_highlight_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            options_frame, text="Surbrillance export", variable=self.list_highlight_var
        ).pack(side="left")

        # Options d'export
        export_frame = ctk.CTkFrame(section3, fg_color="transparent")
        export_frame.pack(fill="x", padx=15, pady=(0, 15))

        self.add_match_col_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            export_frame, text="Ajouter colonne 'Matches'", variable=self.add_match_col_var
        ).pack(side="left", padx=(0, 15))

        self.add_summary_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            export_frame, text="Créer feuille résumé", variable=self.add_summary_var
        ).pack(side="left")

        # === Section Actions ===
        action_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        action_frame.pack(fill="x", pady=(0, 10))

        self.list_search_btn = ctk.CTkButton(
            action_frame,
            text="🔍 RECHERCHER LA LISTE",
            font=("Segoe UI", 14, "bold"),
            height=50,
            fg_color=COLORS["accent_primary"],
            command=self._start_list_search
        )
        self.list_search_btn.pack(fill="x", pady=(0, 10))

        # Progression
        self.list_progress_bar = ctk.CTkProgressBar(action_frame, height=8)
        self.list_progress_bar.pack(fill="x")
        self.list_progress_bar.set(0)

        self.list_status_label = ctk.CTkLabel(
            action_frame, text="Prêt", font=("Segoe UI", 10), text_color=COLORS["text_muted"]
        )
        self.list_status_label.pack(anchor="w", pady=(5, 0))

    def _create_results_tab(self):
        """Onglet des résultats"""
        tab = self.tabview.tab("Résultats")

        # Statistiques
        self.stats_cards = StatCardGroup(tab)
        self.stats_cards.pack(fill="x", padx=10, pady=10)

        self.stats_cards.add_card("total", "Total lignes", "0", "📊")
        self.stats_cards.add_card("matches", "Correspondances", "0", "✓", color=COLORS["success"])
        self.stats_cards.add_card("terms", "Termes trouvés", "0", "🔤", color=COLORS["info"])
        self.stats_cards.add_card("rate", "Taux", "0%", "📈", color=COLORS["warning"])

        # Table de résultats
        self.results_table = PreviewTable(
            tab, title="Résultats de la recherche", max_rows=1000
        )
        self.results_table.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Boutons d'export
        export_frame = ctk.CTkFrame(tab, fg_color="transparent")
        export_frame.pack(fill="x", padx=10, pady=(0, 10))

        self.export_btn = ctk.CTkButton(
            export_frame, text="📤 Exporter les résultats", width=180,
            fg_color=COLORS["success"], state="disabled",
            command=self._export_results
        )
        self.export_btn.pack(side="left", padx=(0, 10))

        self.export_stats_btn = ctk.CTkButton(
            export_frame, text="📊 Exporter avec statistiques", width=180,
            fg_color=COLORS["accent_primary"], state="disabled",
            command=self._export_with_stats
        )
        self.export_stats_btn.pack(side="left")

    def _update_fuzzy_label(self, value):
        self.fuzzy_label.configure(text=f"{int(value)}%")

    def _on_file_loaded(self, df: pd.DataFrame):
        """Callback fichier chargé (recherche simple)"""
        columns = list(df.columns)

        for widget in self.columns_frame.winfo_children():
            widget.destroy()

        self.column_checkboxes = {}
        for col in columns:
            var = ctk.BooleanVar(value=True)
            cb = ctk.CTkCheckBox(
                self.columns_frame, text=col[:40], variable=var, font=("Segoe UI", 10)
            )
            cb.pack(anchor="w", pady=1)
            self.column_checkboxes[col] = var

        self.log_info(f"Fichier chargé: {len(df)} lignes, {len(columns)} colonnes")

    def _on_list_file_loaded(self, df: pd.DataFrame):
        """Callback fichier chargé (recherche par liste)"""
        columns = list(df.columns)
        self.search_col_combo.configure(state="normal", values=columns)
        if columns:
            self.search_col_combo.set(columns[0])

    def _select_all_columns(self, select: bool):
        """Sélectionne/désélectionne toutes les colonnes"""
        for var in self.column_checkboxes.values():
            var.set(select)

    def _toggle_list_mode(self):
        """Bascule entre saisie manuelle et fichier"""
        if self.list_mode_var.get() == "text":
            self.words_text.pack(fill="x", padx=15, pady=(0, 10))
            self.words_file_frame.pack_forget()
        else:
            self.words_text.pack_forget()
            self.words_file_frame.pack(fill="x", padx=15, pady=(0, 15))

    def _browse_words_file(self):
        """Sélectionne un fichier de mots"""
        filepath = filedialog.askopenfilename(
            title="Sélectionner fichier de mots",
            filetypes=[
                ("Fichiers texte", "*.txt"),
                ("Fichiers Excel", "*.xlsx *.xls"),
                ("Tous les fichiers", "*.*")
            ]
        )
        if filepath:
            self.words_file_var.set(filepath)

    def _get_words_list(self) -> List[str]:
        """Récupère la liste de mots selon le mode"""
        if self.list_mode_var.get() == "text":
            text = self.words_text.get("1.0", "end")
            lines = text.strip().split('\n')
            return [l.strip() for l in lines if l.strip() and not l.strip().startswith('#')]
        else:
            filepath = self.words_file_var.get()
            if not filepath:
                return []

            from pathlib import Path
            ext = Path(filepath).suffix.lower()

            if ext in ['.xlsx', '.xls']:
                df = pd.read_excel(filepath)
                return df.iloc[:, 0].dropna().astype(str).tolist()
            else:
                with open(filepath, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                return [l.strip() for l in lines if l.strip() and not l.strip().startswith('#')]

    def _start_simple_search(self):
        """Démarre la recherche simple"""
        if not self.file_selector.is_loaded():
            messagebox.showwarning("Attention", "Veuillez charger un fichier")
            return

        search_text = self.search_entry.get().strip()
        if not search_text:
            messagebox.showwarning("Attention", "Veuillez entrer un terme de recherche")
            return

        self._stop_event.clear()
        self.search_btn.configure(state="disabled")
        self.progress_bar.set(0)

        thread = threading.Thread(target=self._do_simple_search, daemon=True)
        thread.start()

    def _do_simple_search(self):
        """Effectue la recherche simple (dans un thread)"""
        try:
            df = self.file_selector.get_dataframe()
            search_text = self.search_entry.get().strip()
            mode = self.search_mode_var.get()
            case_sensitive = self.case_sensitive_var.get()
            and_mode = self.and_mode_var.get()
            fuzzy_threshold = self.fuzzy_slider.get() / 100

            selected_cols = [col for col, var in self.column_checkboxes.items() if var.get()]
            if not selected_cols:
                selected_cols = list(df.columns)

            terms = [t.strip() for t in search_text.split(',') if t.strip()]

            self.frame.after(0, lambda: self.status_label.configure(text="Recherche en cours..."))

            results = self._search_data(
                df, terms, selected_cols, mode, case_sensitive, and_mode, fuzzy_threshold
            )

            self.df_results = results
            self.frame.after(0, lambda: self._display_results(df, results, terms))

        except Exception as e:
            self.frame.after(0, lambda: messagebox.showerror("Erreur", str(e)))

        finally:
            self.frame.after(0, lambda: self.search_btn.configure(state="normal"))
            self.frame.after(0, lambda: self.progress_bar.set(1.0))

    def _start_list_search(self):
        """Démarre la recherche par liste"""
        if not self.list_file_selector.is_loaded():
            messagebox.showwarning("Attention", "Veuillez charger un fichier Excel")
            return

        words = self._get_words_list()
        if not words:
            messagebox.showwarning("Attention", "Aucun mot à rechercher")
            return

        self._stop_event.clear()
        self.list_search_btn.configure(state="disabled")
        self.list_progress_bar.set(0)

        thread = threading.Thread(target=self._do_list_search, args=(words,), daemon=True)
        thread.start()

    def _do_list_search(self, words: List[str]):
        """Effectue la recherche par liste (dans un thread)"""
        try:
            df = self.list_file_selector.get_dataframe()
            col = self.search_col_combo.get()
            case_sensitive = self.list_case_var.get()
            exact = self.list_exact_var.get()

            self.frame.after(0, lambda: self.list_status_label.configure(text="Recherche en cours..."))

            # Recherche
            found_words = set()
            match_counts = {}

            def check_match(value, word):
                if pd.isna(value):
                    return False
                val = str(value)
                w = word

                if not case_sensitive:
                    val = val.lower()
                    w = w.lower()

                if exact:
                    pattern = r'\b' + re.escape(w) + r'\b'
                    return bool(re.search(pattern, val, 0 if case_sensitive else re.IGNORECASE))
                else:
                    return w in val

            # Créer un masque pour les lignes qui correspondent
            mask = pd.Series([False] * len(df))
            matches_column = []

            for idx, row in df.iterrows():
                if self._stop_event.is_set():
                    break

                cell_value = row[col]
                row_matches = []

                for word in words:
                    if check_match(cell_value, word):
                        row_matches.append(word)
                        found_words.add(word)
                        match_counts[word] = match_counts.get(word, 0) + 1

                if row_matches:
                    mask[idx] = True
                    matches_column.append(", ".join(row_matches))
                else:
                    matches_column.append("")

                # Mise à jour progression
                progress = (idx + 1) / len(df)
                if idx % 100 == 0:
                    self.frame.after(0, lambda p=progress: self.list_progress_bar.set(p))

            # Créer le DataFrame résultat
            results = df[mask].copy()

            if self.add_match_col_var.get():
                results['Matches'] = [m for i, m in enumerate(matches_column) if mask[i]]

            self.df_results = results
            self.match_details = [
                {"word": w, "count": match_counts.get(w, 0)}
                for w in words
            ]

            self.frame.after(0, lambda: self._display_list_results(
                df, results, words, found_words, match_counts
            ))

        except Exception as e:
            self.frame.after(0, lambda: messagebox.showerror("Erreur", str(e)))

        finally:
            self.frame.after(0, lambda: self.list_search_btn.configure(state="normal"))
            self.frame.after(0, lambda: self.list_progress_bar.set(1.0))

    def _search_data(
        self, df: pd.DataFrame, terms: List[str], columns: List[str],
        mode: str, case_sensitive: bool, and_mode: bool, fuzzy_threshold: float
    ) -> pd.DataFrame:
        """Effectue la recherche dans les données"""
        if not case_sensitive:
            terms_search = [t.lower() for t in terms]
        else:
            terms_search = terms

        def match_term(value, term: str) -> bool:
            if pd.isna(value):
                return False

            val = str(value)
            if not case_sensitive:
                val = val.lower()

            if mode == "contains":
                return term in val
            elif mode == "exact":
                pattern = r'\b' + re.escape(term) + r'\b'
                flags = 0 if case_sensitive else re.IGNORECASE
                return bool(re.search(pattern, str(value), flags))
            elif mode == "starts":
                return val.startswith(term)
            elif mode == "ends":
                return val.endswith(term)
            elif mode == "regex":
                try:
                    flags = 0 if case_sensitive else re.IGNORECASE
                    return bool(re.search(term, str(value), flags))
                except re.error:
                    return False
            elif mode == "fuzzy":
                from difflib import SequenceMatcher
                ratio = SequenceMatcher(None, term, val).ratio()
                return ratio >= fuzzy_threshold
            return False

        def row_matches(row) -> bool:
            term_matches = []

            for term in terms_search:
                found = False
                for col in columns:
                    if col in row.index and match_term(row[col], term):
                        found = True
                        break
                term_matches.append(found)

            return all(term_matches) if and_mode else any(term_matches)

        mask = df.apply(row_matches, axis=1)
        return df[mask]

    def _display_results(self, df_original: pd.DataFrame, df_results: pd.DataFrame, terms: List[str]):
        """Affiche les résultats de recherche simple"""
        self.results_table.load_data(df_results)

        total = len(df_original)
        matches = len(df_results)
        rate = (matches / total * 100) if total > 0 else 0

        self.stats_cards.update_card("total", str(total))
        self.stats_cards.update_card("matches", str(matches))
        self.stats_cards.update_card("terms", str(len(terms)))
        self.stats_cards.update_card("rate", f"{rate:.1f}%")

        self.export_btn.configure(state="normal" if matches > 0 else "disabled")
        self.export_stats_btn.configure(state="normal" if matches > 0 else "disabled")

        self.tabview.set("Résultats")
        self.status_label.configure(text=f"{matches} résultat(s) trouvé(s)")
        self.log_success(f"Recherche terminée: {matches}/{total} correspondances")

    def _display_list_results(
        self, df_original: pd.DataFrame, df_results: pd.DataFrame,
        words: List[str], found_words: set, match_counts: Dict
    ):
        """Affiche les résultats de recherche par liste"""
        self.results_table.load_data(df_results)

        total = len(df_original)
        matches = len(df_results)
        terms_found = len(found_words)
        rate = (matches / total * 100) if total > 0 else 0

        self.stats_cards.update_card("total", str(total))
        self.stats_cards.update_card("matches", str(matches))
        self.stats_cards.update_card("terms", f"{terms_found}/{len(words)}")
        self.stats_cards.update_card("rate", f"{rate:.1f}%")

        self.export_btn.configure(state="normal" if matches > 0 else "disabled")
        self.export_stats_btn.configure(state="normal" if matches > 0 else "disabled")

        self.tabview.set("Résultats")
        self.list_status_label.configure(text=f"{matches} résultat(s), {terms_found} termes trouvés")
        self.log_success(f"Recherche terminée: {matches} lignes, {terms_found}/{len(words)} termes trouvés")

    def _export_results(self):
        """Exporte les résultats vers Excel"""
        if self.df_results is None or len(self.df_results) == 0:
            messagebox.showwarning("Attention", "Aucun résultat à exporter")
            return

        filepath = filedialog.asksaveasfilename(
            title="Enregistrer les résultats",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")]
        )

        if filepath:
            success, error = ExcelUtils.write_dataframe_to_excel(
                self.df_results, filepath, "Résultats"
            )

            if success:
                messagebox.showinfo("Succès", f"Résultats exportés vers:\n{filepath}")
                self.log_success(f"Export réussi: {filepath}")
            else:
                messagebox.showerror("Erreur", f"Erreur d'export: {error}")

    def _export_with_stats(self):
        """Exporte les résultats avec statistiques"""
        if self.df_results is None or len(self.df_results) == 0:
            messagebox.showwarning("Attention", "Aucun résultat à exporter")
            return

        filepath = filedialog.asksaveasfilename(
            title="Enregistrer avec statistiques",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")]
        )

        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()

            # Feuille Résultats
            ws_results = wb.active
            ws_results.title = "Résultats"

            header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for r_idx, row in enumerate(dataframe_to_rows(self.df_results, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_results.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font

            # Feuille Statistiques
            if self.match_details:
                ws_stats = wb.create_sheet("Statistiques")

                ws_stats['A1'] = "Terme"
                ws_stats['B1'] = "Occurrences"
                ws_stats['A1'].fill = header_fill
                ws_stats['A1'].font = header_font
                ws_stats['B1'].fill = header_fill
                ws_stats['B1'].font = header_font

                for idx, detail in enumerate(self.match_details, 2):
                    ws_stats.cell(row=idx, column=1, value=detail['word']).border = border
                    ws_stats.cell(row=idx, column=2, value=detail['count']).border = border

                ws_stats.column_dimensions['A'].width = 30
                ws_stats.column_dimensions['B'].width = 15

            wb.save(filepath)
            messagebox.showinfo("Succès", f"Export avec statistiques:\n{filepath}")
            self.log_success(f"Export avec statistiques: {filepath}")

        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def validate_inputs(self) -> tuple[bool, str]:
        return True, ""

    def _execute_task(self) -> Dict[str, Any]:
        return {}

    def update_status(self, message: str, level: str = "info"):
        color = {"info": COLORS["text_muted"], "success": COLORS["success"],
                 "error": COLORS["error"]}.get(level, COLORS["text_muted"])
        self.status_label.configure(text=message, text_color=color)

    def update_progress(self, progress: float):
        self.progress_bar.set(progress)

    def reset(self):
        super().reset()
        self.file_selector.reset()
        self.list_file_selector.reset()
        self.results_table.clear()
        self.stats_cards.reset_all()
        self.search_entry.delete(0, "end")
        self.export_btn.configure(state="disabled")
        self.export_stats_btn.configure(state="disabled")
        self.df_results = None
        self.match_details = []
