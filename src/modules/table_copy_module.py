"""
Module de copie de tableaux Excel
Copie de tableaux entiers avec création de tableaux Excel natifs
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
from datetime import datetime
import threading
import shutil

from .base_module import BaseModule
from ..ui.components.tooltip import Tooltip
from ..ui.components.preview_table import PreviewTable
from ..ui.components.stat_card import StatCardGroup
from ..core.constants import COLORS


class TableCopyModule(BaseModule):
    """
    Module de copie de tableaux Excel

    Fonctionnalités:
    - Copie de tableaux entiers basés sur les en-têtes
    - Création de tableaux Excel natifs avec filtres
    - Traitement par lots avec organisation automatique
    - Réorganisation des colonnes
    """

    MODULE_ID = "table_copy"
    MODULE_NAME = "Copie de tableaux"
    MODULE_DESCRIPTION = "Copie des tableaux Excel avec création de tables natives"
    MODULE_ICON = "📊"

    def __init__(self, *args, **kwargs):
        self.fields: List[Dict[str, str]] = []
        self.files: List[Path] = []
        self.available_sheets: List[str] = []
        super().__init__(*args, **kwargs)

    def _create_interface(self):
        """Crée l'interface du module"""
        main_scroll = ctk.CTkScrollableFrame(self.frame, fg_color="transparent")
        main_scroll.pack(fill="both", expand=True, padx=10, pady=10)

        left_panel = ctk.CTkFrame(main_scroll, fg_color="transparent", width=480)
        left_panel.pack(side="left", fill="y", padx=(0, 10))

        right_panel = ctk.CTkFrame(main_scroll, fg_color="transparent")
        right_panel.pack(side="right", fill="both", expand=True)

        self._create_files_section(left_panel)
        self._create_sheets_section(left_panel)
        self._create_fields_section(left_panel)
        self._create_action_section(left_panel)

        self._create_preview_section(right_panel)

    def _create_files_section(self, parent):
        """Section de sélection des fichiers"""
        section = ctk.CTkFrame(parent, fg_color=COLORS["bg_card"], corner_radius=10)
        section.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section,
            text="📁 FICHIERS À TRAITER",
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Mode de sélection
        mode_frame = ctk.CTkFrame(section, fg_color="transparent")
        mode_frame.pack(fill="x", padx=15, pady=(0, 5))

        ctk.CTkButton(
            mode_frame,
            text="📄 Fichier unique",
            command=self._select_file,
            width=140
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            mode_frame,
            text="📁 Dossier (lot)",
            command=self._select_folder,
            width=140
        ).pack(side="left")

        self.files_label = ctk.CTkLabel(
            section,
            text="Aucun fichier sélectionné",
            text_color=COLORS["text_muted"]
        )
        self.files_label.pack(anchor="w", padx=15, pady=(5, 15))

    def _create_sheets_section(self, parent):
        """Section configuration des feuilles"""
        section = ctk.CTkFrame(parent, fg_color=COLORS["bg_card"], corner_radius=10)
        section.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section,
            text="📋 CONFIGURATION DES FEUILLES",
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 10))

        # Bouton détection
        detect_frame = ctk.CTkFrame(section, fg_color="transparent")
        detect_frame.pack(fill="x", padx=15, pady=(0, 10))

        self.detect_btn = ctk.CTkButton(
            detect_frame,
            text="🔍 Détecter les feuilles",
            command=self._detect_sheets,
            fg_color=COLORS["warning"],
            hover_color="#d35400"
        )
        self.detect_btn.pack(side="left")
        Tooltip(self.detect_btn, "Détecte les feuilles du premier fichier sélectionné")

        self.sheets_status = ctk.CTkLabel(
            detect_frame,
            text="",
            text_color=COLORS["text_muted"]
        )
        self.sheets_status.pack(side="left", padx=10)

        # Feuille source
        source_frame = ctk.CTkFrame(section, fg_color="transparent")
        source_frame.pack(fill="x", padx=15, pady=5)

        ctk.CTkLabel(source_frame, text="Feuille source:", width=120).pack(side="left")

        self.sheet_source_combo = ctk.CTkComboBox(
            source_frame,
            values=["(Détecter les feuilles)"],
            state="disabled",
            width=200,
            command=self._on_sheet_selected
        )
        self.sheet_source_combo.pack(side="left", padx=(10, 0))

        # Feuille cible
        target_frame = ctk.CTkFrame(section, fg_color="transparent")
        target_frame.pack(fill="x", padx=15, pady=(5, 15))

        ctk.CTkLabel(target_frame, text="Feuille cible:", width=120).pack(side="left")

        self.sheet_target_entry = ctk.CTkEntry(target_frame, width=200)
        self.sheet_target_entry.insert(0, "Données Copiées")
        self.sheet_target_entry.pack(side="left", padx=(10, 0))
        Tooltip(self.sheet_target_entry, "Nom de la nouvelle feuille avec le tableau copié")

    def _create_fields_section(self, parent):
        """Section de définition des champs/en-têtes"""
        section = ctk.CTkFrame(parent, fg_color=COLORS["bg_card"], corner_radius=10)
        section.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            section,
            text="🎯 EN-TÊTES DU TABLEAU",
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", padx=15, pady=(15, 5))

        ctk.CTkLabel(
            section,
            text="Définissez les colonnes à copier et leur ordre",
            text_color=COLORS["text_muted"],
            font=("Segoe UI", 10)
        ).pack(anchor="w", padx=15, pady=(0, 10))

        # Zone d'ajout
        add_frame = ctk.CTkFrame(section, fg_color="transparent")
        add_frame.pack(fill="x", padx=15, pady=(0, 10))

        # Nom affiché
        name_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        name_frame.pack(fill="x", pady=2)

        ctk.CTkLabel(name_frame, text="Nom affiché:", width=100, font=("Segoe UI", 10, "bold")).pack(side="left")
        self.field_name_entry = ctk.CTkEntry(name_frame, width=180, placeholder_text="Ex: Référence")
        self.field_name_entry.pack(side="left", padx=(5, 10))

        # Terme à rechercher
        term_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        term_frame.pack(fill="x", pady=2)

        ctk.CTkLabel(term_frame, text="Terme recherché:", width=100, font=("Segoe UI", 10, "bold")).pack(side="left")
        self.field_term_entry = ctk.CTkEntry(term_frame, width=180, placeholder_text="Ex: REF")
        self.field_term_entry.pack(side="left", padx=(5, 10))

        ctk.CTkButton(
            term_frame,
            text="➕",
            width=40,
            fg_color=COLORS["success"],
            command=self._add_field
        ).pack(side="left")

        # Liste des champs
        self.fields_list_frame = ctk.CTkScrollableFrame(section, fg_color="transparent", height=160)
        self.fields_list_frame.pack(fill="x", padx=15, pady=(0, 10))

        self.fields_count_label = ctk.CTkLabel(
            section,
            text="Aucun en-tête défini",
            text_color=COLORS["warning"]
        )
        self.fields_count_label.pack(anchor="w", padx=15, pady=(0, 15))

        self._update_fields_list()

    def _create_action_section(self, parent):
        """Section des actions"""
        section = ctk.CTkFrame(parent, fg_color="transparent")
        section.pack(fill="x", pady=(0, 10))

        # Option déplacement auto
        self.auto_move_var = ctk.BooleanVar(value=True)
        auto_cb = ctk.CTkCheckBox(
            section,
            text="Déplacer fichiers vers Traités/Non Traités",
            variable=self.auto_move_var
        )
        auto_cb.pack(anchor="w", pady=(0, 10))
        Tooltip(auto_cb, "Organise automatiquement les fichiers après traitement")

        # Option tableau natif
        self.native_table_var = ctk.BooleanVar(value=True)
        native_cb = ctk.CTkCheckBox(
            section,
            text="Créer tableau Excel natif avec filtres",
            variable=self.native_table_var
        )
        native_cb.pack(anchor="w", pady=(0, 10))
        Tooltip(native_cb, "Crée un vrai tableau Excel avec filtres automatiques")

        # Boutons
        self.process_btn = ctk.CTkButton(
            section,
            text="🚀 COPIER LES TABLEAUX",
            font=("Segoe UI", 14, "bold"),
            height=50,
            fg_color=COLORS["success"],
            command=self.start_execution
        )
        self.process_btn.pack(fill="x", pady=(0, 10))

        # Progression
        self.progress_bar = ctk.CTkProgressBar(section, height=8)
        self.progress_bar.pack(fill="x")
        self.progress_bar.set(0)

        self.status_label = ctk.CTkLabel(
            section,
            text="Prêt",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"]
        )
        self.status_label.pack(anchor="w", pady=(5, 0))

    def _create_preview_section(self, parent):
        """Section de prévisualisation"""
        # Statistiques
        self.stats_cards = StatCardGroup(parent)
        self.stats_cards.pack(fill="x", pady=(0, 10))

        self.stats_cards.add_card("total", "Fichiers", "0", "📁")
        self.stats_cards.add_card("success", "Copiés", "0", "✓", color=COLORS["success"])
        self.stats_cards.add_card("rows", "Lignes", "0", "📝", color=COLORS["info"])
        self.stats_cards.add_card("errors", "Erreurs", "0", "✗", color=COLORS["error"])

        # Aperçu
        self.preview_table = PreviewTable(
            parent,
            title="Aperçu du tableau source"
        )
        self.preview_table.pack(fill="both", expand=True)

    def _select_file(self):
        """Sélectionne un fichier unique"""
        filepath = filedialog.askopenfilename(
            title="Sélectionner un fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls *.xlsm")]
        )
        if filepath:
            self.files = [Path(filepath)]
            self._update_files_label()
            self.log_info(f"Fichier sélectionné: {Path(filepath).name}")

    def _select_folder(self):
        """Sélectionne un dossier pour traitement par lot"""
        folder = filedialog.askdirectory(title="Sélectionner le dossier")
        if folder:
            folder_path = Path(folder)
            self.files = list(folder_path.glob("*.xlsx")) + \
                         list(folder_path.glob("*.xls")) + \
                         list(folder_path.glob("*.xlsm"))
            # Exclure les sous-dossiers
            self.files = [f for f in self.files if f.parent == folder_path]
            self._update_files_label()
            self.log_info(f"Dossier sélectionné: {len(self.files)} fichiers")

    def _update_files_label(self):
        """Met à jour le label du nombre de fichiers"""
        n = len(self.files)
        if n == 0:
            self.files_label.configure(
                text="Aucun fichier sélectionné",
                text_color=COLORS["text_muted"]
            )
        elif n == 1:
            self.files_label.configure(
                text=f"✓ {self.files[0].name}",
                text_color=COLORS["success"]
            )
        else:
            self.files_label.configure(
                text=f"✓ {n} fichiers sélectionnés",
                text_color=COLORS["success"]
            )

    def _detect_sheets(self):
        """Détecte les feuilles du premier fichier"""
        if not self.files:
            messagebox.showwarning("Attention", "Veuillez d'abord sélectionner un fichier")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.files[0], read_only=True)
            self.available_sheets = wb.sheetnames
            wb.close()

            if self.available_sheets:
                self.sheet_source_combo.configure(
                    state="normal",
                    values=self.available_sheets
                )
                self.sheet_source_combo.set(self.available_sheets[0])
                self.sheets_status.configure(
                    text=f"✓ {len(self.available_sheets)} feuille(s)",
                    text_color=COLORS["success"]
                )
                self.log_success(f"Feuilles détectées: {', '.join(self.available_sheets)}")
            else:
                self.sheets_status.configure(
                    text="✗ Aucune feuille",
                    text_color=COLORS["error"]
                )

        except Exception as e:
            self.log_error(f"Erreur détection feuilles: {e}")
            messagebox.showerror("Erreur", f"Impossible de lire le fichier:\n{e}")

    def _on_sheet_selected(self, sheet_name: str):
        """Callback quand une feuille est sélectionnée"""
        self.log_info(f"Feuille source: {sheet_name}")

    def _add_field(self):
        """Ajoute un champ/en-tête"""
        name = self.field_name_entry.get().strip()
        term = self.field_term_entry.get().strip()

        if not name or not term:
            messagebox.showwarning("Attention", "Veuillez remplir le nom et le terme")
            return

        # Vérifier doublon
        if any(f['nom'] == name for f in self.fields):
            messagebox.showwarning("Attention", f"Le champ '{name}' existe déjà")
            return

        self.fields.append({"nom": name, "terme_recherche": term})
        self.field_name_entry.delete(0, "end")
        self.field_term_entry.delete(0, "end")
        self.field_name_entry.focus()

        self._update_fields_list()
        self.log_info(f"Champ ajouté: {name} → {term}")

    def _remove_field(self, index: int):
        """Supprime un champ"""
        if 0 <= index < len(self.fields):
            del self.fields[index]
            self._update_fields_list()

    def _move_field(self, index: int, direction: int):
        """Déplace un champ vers le haut (-1) ou le bas (+1)"""
        new_index = index + direction
        if 0 <= new_index < len(self.fields):
            self.fields[index], self.fields[new_index] = self.fields[new_index], self.fields[index]
            self._update_fields_list()

    def _update_fields_list(self):
        """Met à jour l'affichage de la liste des champs"""
        for widget in self.fields_list_frame.winfo_children():
            widget.destroy()

        if not self.fields:
            ctk.CTkLabel(
                self.fields_list_frame,
                text="Aucun en-tête défini. Ajoutez-en ci-dessus.",
                text_color=COLORS["text_muted"]
            ).pack(pady=10)
            self.fields_count_label.configure(
                text="⚠ Aucun en-tête défini",
                text_color=COLORS["warning"]
            )
            return

        for idx, field in enumerate(self.fields):
            frame = ctk.CTkFrame(self.fields_list_frame, fg_color=("gray85", "gray25"))
            frame.pack(fill="x", pady=2)

            # Numéro et info
            ctk.CTkLabel(
                frame,
                text=f"{idx + 1}.",
                font=("Segoe UI", 11, "bold"),
                width=25
            ).pack(side="left", padx=5, pady=5)

            info_frame = ctk.CTkFrame(frame, fg_color="transparent")
            info_frame.pack(side="left", fill="x", expand=True, padx=5)

            ctk.CTkLabel(
                info_frame,
                text=field['nom'],
                font=("Segoe UI", 11, "bold")
            ).pack(anchor="w")

            ctk.CTkLabel(
                info_frame,
                text=f"→ '{field['terme_recherche']}'",
                text_color=COLORS["text_muted"],
                font=("Segoe UI", 9)
            ).pack(anchor="w")

            # Boutons
            btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
            btn_frame.pack(side="right", padx=5, pady=3)

            # Monter
            ctk.CTkButton(
                btn_frame,
                text="↑",
                width=28,
                height=24,
                fg_color=COLORS["info"] if idx > 0 else "gray",
                command=lambda i=idx: self._move_field(i, -1),
                state="normal" if idx > 0 else "disabled"
            ).pack(side="left", padx=1)

            # Descendre
            ctk.CTkButton(
                btn_frame,
                text="↓",
                width=28,
                height=24,
                fg_color=COLORS["info"] if idx < len(self.fields) - 1 else "gray",
                command=lambda i=idx: self._move_field(i, 1),
                state="normal" if idx < len(self.fields) - 1 else "disabled"
            ).pack(side="left", padx=1)

            # Supprimer
            ctk.CTkButton(
                btn_frame,
                text="✖",
                width=28,
                height=24,
                fg_color=COLORS["error"],
                command=lambda i=idx: self._remove_field(i)
            ).pack(side="left", padx=1)

        self.fields_count_label.configure(
            text=f"✓ {len(self.fields)} en-tête(s) défini(s)",
            text_color=COLORS["success"]
        )

    def validate_inputs(self) -> tuple[bool, str]:
        """Valide les entrées"""
        if not self.files:
            return False, "Aucun fichier sélectionné"

        if not self.fields:
            return False, "Aucun en-tête défini"

        sheet = self.sheet_source_combo.get()
        if not sheet or sheet == "(Détecter les feuilles)":
            return False, "Veuillez détecter et sélectionner une feuille source"

        return True, ""

    def _execute_task(self) -> Dict[str, Any]:
        """Exécute la copie des tableaux"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        import openpyxl.utils

        sheet_source = self.sheet_source_combo.get()
        sheet_target = self.sheet_target_entry.get().strip() or "Données Copiées"
        auto_move = self.auto_move_var.get()
        create_native = self.native_table_var.get()

        total = len(self.files)
        success = 0
        total_rows = 0
        errors = 0

        for idx, filepath in enumerate(self.files):
            if self.is_cancelled():
                break

            self.update_progress((idx + 1) / total)
            self.update_status(f"Traitement: {filepath.name}")

            try:
                result = self._copy_table_from_file(
                    filepath, sheet_source, sheet_target, create_native
                )

                if result["success"]:
                    success += 1
                    total_rows += result["rows"]
                    self.log_success(f"Copié: {filepath.name} ({result['rows']} lignes)")

                    if auto_move:
                        self._move_file(filepath, True)
                else:
                    errors += 1
                    self.log_warning(f"Échec: {filepath.name} - {result['error']}")

                    if auto_move:
                        self._move_file(filepath, False)

            except Exception as e:
                errors += 1
                self.log_error(f"Erreur {filepath.name}: {e}")

                if auto_move:
                    self._move_file(filepath, False)

        # Mise à jour interface
        self.frame.after(0, lambda: self._update_stats(total, success, total_rows, errors))

        return {"total": total, "success": success, "rows": total_rows, "errors": errors}

    def _copy_table_from_file(self, filepath: Path, sheet_source: str,
                               sheet_target: str, create_native: bool) -> Dict[str, Any]:
        """Copie le tableau d'un fichier"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        import openpyxl.utils

        wb = load_workbook(filepath)

        # Vérifier la feuille source
        if sheet_source not in wb.sheetnames:
            # Essayer la première feuille
            sheet_source = wb.sheetnames[0]

        ws_source = wb[sheet_source]

        # Trouver les en-têtes
        header_info = self._find_headers(ws_source)
        if header_info is None:
            wb.close()
            return {"success": False, "error": "En-têtes non trouvés", "rows": 0}

        header_row, mapping = header_info

        # Trouver la fin des données
        columns = [info['col'] for info in mapping.values()]
        first_col = min(columns)
        last_col = max(columns)
        last_row = self._find_data_end(ws_source, header_row, first_col, last_col)

        nb_rows = last_row - header_row
        if nb_rows <= 0:
            wb.close()
            return {"success": False, "error": "Aucune donnée", "rows": 0}

        # Créer la feuille cible
        if sheet_target in wb.sheetnames:
            del wb[sheet_target]
        ws_target = wb.create_sheet(sheet_target)

        # Copier les données
        for row_offset in range(nb_rows + 1):
            source_row = header_row + row_offset
            target_row = row_offset + 1

            for target_col, field in enumerate(self.fields, 1):
                field_name = field['nom']
                source_col = mapping[field_name]['col']

                source_cell = ws_source.cell(source_row, source_col)
                target_cell = ws_target.cell(target_row, target_col)

                # Première ligne = en-têtes personnalisés
                if row_offset == 0:
                    target_cell.value = field_name
                else:
                    target_cell.value = source_cell.value
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format

        # Créer tableau natif si demandé
        if create_native:
            self._create_native_table(ws_target, nb_rows + 1, len(self.fields), sheet_target)

        # Ajuster les colonnes
        self._adjust_columns(ws_target, len(self.fields))

        # Figer les volets
        ws_target.freeze_panes = 'A2'

        wb.save(filepath)
        wb.close()

        return {"success": True, "error": "", "rows": nb_rows}

    def _find_headers(self, worksheet) -> Optional[Tuple[int, Dict]]:
        """Trouve les en-têtes dans la feuille"""
        for row in range(1, 21):
            mapping = {}

            for col in range(1, 701):
                cell_value = worksheet.cell(row, col).value
                if cell_value:
                    cell_text = str(cell_value).strip().lower()

                    for field in self.fields:
                        term = field['terme_recherche'].strip().lower()
                        if cell_text == term:
                            mapping[field['nom']] = {'col': col, 'terme': field['terme_recherche']}

            if len(mapping) == len(self.fields):
                return row, mapping

        return None

    def _find_data_end(self, worksheet, header_row: int, first_col: int,
                       last_col: int, max_empty: int = 2) -> int:
        """Trouve la dernière ligne de données"""
        empty_count = 0
        last_data_row = header_row

        for row in range(header_row + 1, worksheet.max_row + 1):
            is_empty = True
            for col in range(first_col, last_col + 1):
                value = worksheet.cell(row, col).value
                if value and str(value).strip():
                    is_empty = False
                    break

            if is_empty:
                empty_count += 1
                if empty_count >= max_empty:
                    break
            else:
                empty_count = 0
                last_data_row = row

        return last_data_row

    def _create_native_table(self, worksheet, nb_rows: int, nb_cols: int, name: str):
        """Crée un tableau Excel natif avec filtres"""
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import openpyxl.utils

        try:
            last_col_letter = openpyxl.utils.get_column_letter(nb_cols)
            ref = f"A1:{last_col_letter}{nb_rows}"

            # Nom du tableau sans caractères spéciaux
            table_name = "Tbl_" + "".join(c for c in name if c.isalnum())[:20]

            tab = Table(displayName=table_name, ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            worksheet.add_table(tab)

            # Styles supplémentaires
            font_header = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='medium', color='1F4E78'),
                right=Side(style='medium', color='1F4E78'),
                top=Side(style='medium', color='1F4E78'),
                bottom=Side(style='medium', color='1F4E78')
            )

            for col in range(1, nb_cols + 1):
                cell = worksheet.cell(1, col)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border

            worksheet.row_dimensions[1].height = 30

        except Exception as e:
            self.log_warning(f"Impossible de créer le tableau natif: {e}")

    def _adjust_columns(self, worksheet, nb_cols: int):
        """Ajuste la largeur des colonnes"""
        import openpyxl.utils

        for col in range(1, nb_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            max_length = 0

            for row in range(1, min(worksheet.max_row + 1, 100)):
                cell = worksheet.cell(row, col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            width = min(max(max_length + 2, 12), 50)
            worksheet.column_dimensions[col_letter].width = width

    def _move_file(self, filepath: Path, success: bool):
        """Déplace le fichier vers Traités ou Non Traités"""
        try:
            parent = filepath.parent
            dest_folder = parent / ("Traités" if success else "Non Traités")
            dest_folder.mkdir(exist_ok=True)

            destination = dest_folder / filepath.name

            # Gérer les conflits de nom
            counter = 1
            while destination.exists():
                stem = filepath.stem
                suffix = filepath.suffix
                destination = dest_folder / f"{stem}_{counter}{suffix}"
                counter += 1

            shutil.move(str(filepath), str(destination))

        except Exception as e:
            self.log_warning(f"Impossible de déplacer {filepath.name}: {e}")

    def _update_stats(self, total: int, success: int, rows: int, errors: int):
        """Met à jour les statistiques"""
        self.stats_cards.update_card("total", str(total))
        self.stats_cards.update_card("success", str(success))
        self.stats_cards.update_card("rows", str(rows))
        self.stats_cards.update_card("errors", str(errors))

        if errors > 0:
            self.update_status(f"Terminé avec {errors} erreur(s)", "warning")
        else:
            self.update_status(f"Terminé: {success}/{total} fichiers, {rows} lignes copiées", "success")

    def update_status(self, message: str, level: str = "info"):
        color_map = {
            "info": COLORS["text_muted"],
            "success": COLORS["success"],
            "warning": COLORS["warning"],
            "error": COLORS["error"]
        }
        self.status_label.configure(
            text=message,
            text_color=color_map.get(level, COLORS["text_muted"])
        )

    def update_progress(self, progress: float):
        self.progress_bar.set(progress)

    def reset(self):
        super().reset()
        self.files = []
        self.fields = []
        self.available_sheets = []
        self._update_files_label()
        self._update_fields_list()
        self.preview_table.clear()
        self.stats_cards.reset_all()
        self.sheet_source_combo.configure(values=["(Détecter les feuilles)"], state="disabled")
        self.sheets_status.configure(text="")
