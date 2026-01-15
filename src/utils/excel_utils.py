"""
Utilitaires Excel centralisés pour ExcelToolsPro
Fonctions communes de lecture/écriture/formatage Excel
Utilise la configuration centralisée pour tous les paramètres
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


class ExcelUtils:
    """Classe utilitaire pour les opérations Excel"""

    # Styles par défaut (utilisés si pas de config fournie)
    DEFAULT_HEADER_COLOR = "1F4E79"
    DEFAULT_HEADER_FONT_COLOR = "FFFFFF"
    DEFAULT_ALTERNATE_COLOR = "F2F2F2"
    DEFAULT_SUCCESS_COLOR = "C6EFCE"
    DEFAULT_ERROR_COLOR = "FFC7CE"
    DEFAULT_WARNING_COLOR = "FFEB9C"

    @staticmethod
    def _hex_to_rgb(hex_color: str) -> str:
        """Convertit une couleur hex (#RRGGBB) en format openpyxl (RRGGBB)"""
        return hex_color.lstrip('#').upper()

    @staticmethod
    def read_excel_file(
        filepath: str,
        sheet_name: Optional[str] = None,
        as_string: bool = True
    ) -> Tuple[Optional[pd.DataFrame], List[str], Optional[str]]:
        """
        Lit un fichier Excel et retourne les données

        Args:
            filepath: Chemin du fichier Excel
            sheet_name: Nom de l'onglet à lire (None = premier onglet)
            as_string: Convertir toutes les données en string

        Returns:
            Tuple (DataFrame, liste des onglets, message d'erreur ou None)
        """
        try:
            xl = pd.ExcelFile(filepath)
            sheets = xl.sheet_names

            if sheet_name is None:
                sheet_name = sheets[0]
            elif sheet_name not in sheets:
                return None, sheets, f"Onglet '{sheet_name}' introuvable"

            dtype = str if as_string else None
            df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=dtype)
            df.columns = df.columns.str.strip()

            return df, sheets, None

        except Exception as e:
            return None, [], str(e)

    @staticmethod
    def get_excel_sheets(filepath: str) -> List[str]:
        """Récupère la liste des feuilles d'un fichier Excel"""
        try:
            xl = pd.ExcelFile(filepath)
            return xl.sheet_names
        except Exception:
            return []

    @staticmethod
    def get_sheet_names(filepath: str) -> Tuple[List[str], Optional[str]]:
        """
        Récupère la liste des onglets d'un fichier Excel

        Returns:
            Tuple (liste des onglets, message d'erreur ou None)
        """
        try:
            xl = pd.ExcelFile(filepath)
            return xl.sheet_names, None
        except Exception as e:
            return [], str(e)

    @staticmethod
    def write_dataframe_to_excel(
        df: pd.DataFrame,
        filepath: str,
        sheet_name: str = "Sheet1",
        apply_formatting: bool = True,
        freeze_header: bool = True,
        auto_fit_columns: bool = True,
        alternate_rows: bool = True,
        add_borders: bool = True,
        header_bg_color: str = "#1F4E79",
        header_font_color: str = "#FFFFFF",
        alternate_row_color: str = "#F2F2F2",
        min_column_width: int = 10,
        max_column_width: int = 50,
        autofit_sample_rows: int = 100
    ) -> Tuple[bool, Optional[str]]:
        """
        Écrit un DataFrame dans un fichier Excel avec formatage professionnel

        Args:
            df: DataFrame à écrire
            filepath: Chemin du fichier de destination
            sheet_name: Nom de l'onglet
            apply_formatting: Appliquer le formatage
            freeze_header: Geler l'en-tête
            auto_fit_columns: Ajuster les colonnes automatiquement
            alternate_rows: Alternance de couleurs
            add_borders: Ajouter des bordures
            header_bg_color: Couleur de fond de l'en-tête (hex)
            header_font_color: Couleur de police de l'en-tête (hex)
            alternate_row_color: Couleur d'alternance (hex)
            min_column_width: Largeur minimale des colonnes
            max_column_width: Largeur maximale des colonnes
            autofit_sample_rows: Nombre de lignes à analyser pour l'auto-fit

        Returns:
            Tuple (succès, message d'erreur ou None)
        """
        try:
            filepath = Path(filepath)

            # Charger ou créer le workbook
            if filepath.exists():
                wb = load_workbook(filepath)
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
            else:
                wb = Workbook()
                # Supprimer la feuille par défaut
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            ws = wb.create_sheet(sheet_name)

            # Préparer les styles
            header_fill = PatternFill(
                start_color=ExcelUtils._hex_to_rgb(header_bg_color),
                end_color=ExcelUtils._hex_to_rgb(header_bg_color),
                fill_type="solid"
            )
            header_font = Font(
                bold=True,
                color=ExcelUtils._hex_to_rgb(header_font_color),
                size=11
            )
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            alternate_fill = PatternFill(
                start_color=ExcelUtils._hex_to_rgb(alternate_row_color),
                end_color=ExcelUtils._hex_to_rgb(alternate_row_color),
                fill_type="solid"
            )

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ) if add_borders else None

            # Écrire les données
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    if apply_formatting:
                        if add_borders:
                            cell.border = thin_border

                        # En-tête
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                        else:
                            cell.alignment = Alignment(vertical='center')
                            # Alternance des lignes
                            if alternate_rows and r_idx % 2 == 0:
                                cell.fill = alternate_fill

            # Ajustement des colonnes
            if auto_fit_columns:
                ExcelUtils._auto_fit_columns(
                    ws, df,
                    min_width=min_column_width,
                    max_width=max_column_width,
                    sample_rows=autofit_sample_rows
                )

            # Gel des volets
            if freeze_header:
                ws.freeze_panes = 'A2'

            wb.save(filepath)
            wb.close()

            return True, None

        except Exception as e:
            return False, str(e)

    @staticmethod
    def write_with_config(
        df: pd.DataFrame,
        filepath: str,
        sheet_name: str,
        config
    ) -> Tuple[bool, Optional[str]]:
        """
        Écrit un DataFrame en utilisant la configuration centralisée

        Args:
            df: DataFrame à écrire
            filepath: Chemin du fichier
            sheet_name: Nom de l'onglet
            config: Instance de ExcelExportConfig

        Returns:
            Tuple (succès, message d'erreur ou None)
        """
        return ExcelUtils.write_dataframe_to_excel(
            df=df,
            filepath=filepath,
            sheet_name=sheet_name,
            apply_formatting=True,
            freeze_header=config.freeze_header,
            auto_fit_columns=config.auto_fit_columns,
            alternate_rows=config.alternate_row_colors,
            add_borders=config.add_borders,
            header_bg_color=config.header_bg_color,
            header_font_color=config.header_font_color,
            alternate_row_color=config.alternate_row_color,
            min_column_width=config.min_column_width,
            max_column_width=config.max_column_width,
            autofit_sample_rows=config.autofit_sample_rows
        )

    @staticmethod
    def _auto_fit_columns(
        ws,
        df: pd.DataFrame,
        min_width: int = 10,
        max_width: int = 50,
        sample_rows: int = 100
    ):
        """Ajuste automatiquement la largeur des colonnes"""
        for col_idx, col_name in enumerate(df.columns, start=1):
            # Calculer la largeur maximale
            max_length = len(str(col_name))

            # Limiter le nombre de lignes analysées pour la performance
            for row_idx in range(2, min(len(df) + 2, sample_rows + 2)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Appliquer avec limites
            width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    @staticmethod
    def add_sheet_to_workbook(
        filepath: str,
        sheet_name: str,
        df: pd.DataFrame,
        apply_formatting: bool = True,
        config=None
    ) -> Tuple[bool, Optional[str]]:
        """
        Ajoute un onglet à un fichier Excel existant

        Args:
            filepath: Chemin du fichier
            sheet_name: Nom de l'onglet
            df: DataFrame à écrire
            apply_formatting: Appliquer le formatage
            config: Instance de ExcelExportConfig (optionnel)

        Returns:
            Tuple (succès, message d'erreur ou None)
        """
        try:
            wb = load_workbook(filepath)

            # Supprimer l'onglet s'il existe
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

            ws = wb.create_sheet(sheet_name)

            # Utiliser la config si fournie
            if config:
                header_fill = PatternFill(
                    start_color=ExcelUtils._hex_to_rgb(config.header_bg_color),
                    end_color=ExcelUtils._hex_to_rgb(config.header_bg_color),
                    fill_type="solid"
                )
                header_font = Font(
                    bold=config.header_font_bold,
                    color=ExcelUtils._hex_to_rgb(config.header_font_color),
                    size=config.header_font_size
                )
            else:
                header_fill = PatternFill(
                    start_color=ExcelUtils.DEFAULT_HEADER_COLOR,
                    end_color=ExcelUtils.DEFAULT_HEADER_COLOR,
                    fill_type="solid"
                )
                header_font = Font(bold=True, color=ExcelUtils.DEFAULT_HEADER_FONT_COLOR, size=11)

            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Écrire les données
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    if apply_formatting:
                        cell.border = thin_border

                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                        else:
                            cell.alignment = Alignment(vertical='center')

            # Ajustement des colonnes
            min_w = config.min_column_width if config else 10
            max_w = config.max_column_width if config else 50
            ExcelUtils._auto_fit_columns(ws, df, min_width=min_w, max_width=max_w)

            # Gel des volets
            freeze = config.freeze_header if config else True
            if freeze:
                ws.freeze_panes = 'A2'

            wb.save(filepath)
            wb.close()

            return True, None

        except Exception as e:
            return False, str(e)

    @staticmethod
    def apply_conditional_formatting(
        ws,
        column_name: str,
        df: pd.DataFrame,
        condition_map: Dict[str, PatternFill]
    ):
        """
        Applique un formatage conditionnel à une colonne

        Args:
            ws: Worksheet openpyxl
            column_name: Nom de la colonne
            df: DataFrame source
            condition_map: Dictionnaire {valeur: PatternFill}
        """
        if column_name not in df.columns:
            return

        col_idx = list(df.columns).index(column_name) + 1

        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = str(cell.value) if cell.value else ""

            for condition_value, fill in condition_map.items():
                if value.upper() == condition_value.upper():
                    cell.fill = fill
                    break

    @staticmethod
    def merge_excel_files(
        file_paths: List[str],
        output_path: str,
        skip_headers: bool = True,
        config=None
    ) -> Tuple[bool, int, Optional[str]]:
        """
        Fusionne plusieurs fichiers Excel en un seul

        Args:
            file_paths: Liste des chemins des fichiers à fusionner
            output_path: Chemin du fichier de sortie
            skip_headers: Ignorer les en-têtes (sauf premier fichier)
            config: Instance de ExcelExportConfig (optionnel)

        Returns:
            Tuple (succès, nombre de lignes, message d'erreur ou None)
        """
        try:
            all_data = []

            for i, filepath in enumerate(file_paths):
                df, _, error = ExcelUtils.read_excel_file(filepath)
                if error:
                    return False, 0, f"Erreur lecture {filepath}: {error}"

                if skip_headers and i > 0:
                    df = df.iloc[1:]

                all_data.append(df)

            merged_df = pd.concat(all_data, ignore_index=True)

            if config:
                success, error = ExcelUtils.write_with_config(
                    merged_df, output_path, "Données_Fusionnées", config
                )
            else:
                success, error = ExcelUtils.write_dataframe_to_excel(
                    merged_df, output_path, "Données_Fusionnées"
                )

            return success, len(merged_df), error

        except Exception as e:
            return False, 0, str(e)

    @staticmethod
    def search_in_excel(
        df: pd.DataFrame,
        search_term: str,
        columns: Optional[List[str]] = None,
        case_sensitive: bool = False,
        exact_match: bool = False
    ) -> pd.DataFrame:
        """
        Recherche un terme dans un DataFrame

        Args:
            df: DataFrame à rechercher
            search_term: Terme à rechercher
            columns: Colonnes à rechercher (None = toutes)
            case_sensitive: Sensibilité à la casse
            exact_match: Correspondance exacte

        Returns:
            DataFrame filtré
        """
        if columns is None:
            columns = df.columns.tolist()

        if not case_sensitive:
            search_term = search_term.lower()

        mask = pd.Series([False] * len(df))

        for col in columns:
            if col not in df.columns:
                continue

            col_values = df[col].astype(str)
            if not case_sensitive:
                col_values = col_values.str.lower()

            if exact_match:
                mask |= (col_values == search_term)
            else:
                mask |= col_values.str.contains(search_term, na=False, regex=False)

        return df[mask]

    @staticmethod
    def get_column_statistics(df: pd.DataFrame, column: str) -> Dict[str, Any]:
        """
        Calcule des statistiques sur une colonne

        Returns:
            Dictionnaire de statistiques
        """
        if column not in df.columns:
            return {}

        col = df[column]
        stats = {
            "total": len(col),
            "non_vides": col.notna().sum(),
            "vides": col.isna().sum(),
            "uniques": col.nunique(),
        }

        # Pour les colonnes numériques
        if pd.api.types.is_numeric_dtype(col):
            stats.update({
                "min": col.min(),
                "max": col.max(),
                "moyenne": col.mean(),
                "somme": col.sum(),
            })

        return stats

    @staticmethod
    def get_status_fills(config=None) -> Dict[str, PatternFill]:
        """
        Retourne les PatternFill pour les différents statuts

        Args:
            config: Instance de ExcelExportConfig (optionnel)

        Returns:
            Dict avec les fills success, error, warning
        """
        if config:
            return {
                "success": PatternFill(
                    start_color=ExcelUtils._hex_to_rgb(config.success_color),
                    end_color=ExcelUtils._hex_to_rgb(config.success_color),
                    fill_type="solid"
                ),
                "error": PatternFill(
                    start_color=ExcelUtils._hex_to_rgb(config.error_color),
                    end_color=ExcelUtils._hex_to_rgb(config.error_color),
                    fill_type="solid"
                ),
                "warning": PatternFill(
                    start_color=ExcelUtils._hex_to_rgb(config.warning_color),
                    end_color=ExcelUtils._hex_to_rgb(config.warning_color),
                    fill_type="solid"
                ),
            }
        else:
            return {
                "success": PatternFill(
                    start_color=ExcelUtils.DEFAULT_SUCCESS_COLOR,
                    end_color=ExcelUtils.DEFAULT_SUCCESS_COLOR,
                    fill_type="solid"
                ),
                "error": PatternFill(
                    start_color=ExcelUtils.DEFAULT_ERROR_COLOR,
                    end_color=ExcelUtils.DEFAULT_ERROR_COLOR,
                    fill_type="solid"
                ),
                "warning": PatternFill(
                    start_color=ExcelUtils.DEFAULT_WARNING_COLOR,
                    end_color=ExcelUtils.DEFAULT_WARNING_COLOR,
                    fill_type="solid"
                ),
            }
